#------------------------#
# 2D-FD ver 1.0          #
# parameters             #
# author:Kazumi Ishiwata #
#------------------------#
import datetime   as dtm
from   matplotlib import cm
import numpy      as np
import openpyxl   as xl
import os         as os
import shutil     as sh

#excelの指定範囲を2次元配列として取得する
def get_list_2d(sheet,start_col,end_col,start_row,end_row)  -> np.ndarray:
    return np.flipud(np.array(get_value_list(sheet.iter_rows(min_row = start_row,
                                                             max_row =   end_row,
                                                             min_col = start_col,
                                                             max_col =   end_col)),dtype=float)).T
def get_value_list(t_2d)  -> np.ndarray:
    return([[cell.value for cell in row] for row in t_2d])

#excelの指定範囲に二次元配列を書き込む
def write_list_2d(sheet,l_2d,start_row,start_col) -> None:
    for x, i in enumerate(l_2d.T):
        for y, j in enumerate(i):
            sheet.cell(row    = start_col + x,
                       column = start_row + y,
                       value  = np.flipud(l_2d.T)[x][y])

#結果をconfig.xlsxに書き込む
def writeResult() -> None:
    print('config.xlsx 書込開始')
    write_list_2d(uxSheet_o,ux        ,1,1)
    write_list_2d(uySheet_o,uy        ,1,1)
    write_list_2d( pSheet_o,p         ,1,1)
    book_o.save(dirName + '\\03_Last\\config.xlsx')
    write_list_2d(mkSheet_o,marker_scr,1,1)
    settings_o['K10'] =   start_time.strftime('%Y/%m/%d  %H:%M:%S') #計算開始時刻
    end_time          = dtm.datetime.now(JST)
    settings_o['K11'] =     end_time.strftime('%Y/%m/%d  %H:%M:%S') #計算完了時刻
    settings_o['K12'] = str(end_time - start_time)                  #所要時間
    settings_o['K13'] = result_all                                  #計算結果
    book_o.save(dirName + '\\03_Last\\config.xlsx')
    print('config.xlsx 書込完了')

#以下設定読込-----------------------------------------------------------------------------

#開始時刻を記録する
t_delta = dtm.timedelta(hours=9)
JST = dtm.timezone(t_delta,'JST')
start_time :dtm.datetime = dtm.datetime.now(JST)

result_all :str
result_all = 'config.xlsx 読込開始'
print(result_all)

#config.xlsx読み込みのための設定
book_i :xl.Workbook = xl.load_workbook('config.xlsx', data_only=True)
settings_i = book_i['settings']
baSheet_i  = book_i['barrier' ]
mkSheet_i  = book_i['marker'  ]
uxSheet_i  = book_i['ux'      ]
uySheet_i  = book_i['uy'      ]
pSheet_i   = book_i['p'       ]
KxSheet_i  = book_i['Kx'      ]
KySheet_i  = book_i['Ky'      ]


#流体の性質
ro :float = settings_i['C4' ].value #[kg/m^3]   密度
mu :float = settings_i['C5' ].value #[kg/(m s)] 粘度
nu :float = mu / ro                 #[m^2/s]    動粘度

#その他設定
drawing_mode    :str   = settings_i['C11'].value           #markerの描画方法
animation_speed :str   = settings_i['C12'].value           #アニメーションの速さ(normal/slow)
random_marker   :bool  = settings_i['C13'].value == 'true' #ランダムな配置のマーカーを追加するか
folder_name     :str   = settings_i['C14'].value           #計算結果を格納するフォルダの名前
command_stop    :bool  = settings_i['C15'].value == 'true' #[E]長押しで計算を途中で止めるか 

#計算領域(空間)
x_min :float = settings_i['G4' ].value #[m]  left
x_max :float = settings_i['G5' ].value #[m]  right
y_min :float = settings_i['G6' ].value #[m]  bottom
y_max :float = settings_i['G7' ].value #[m]  up
dx    :float = settings_i['G8' ].value #[m]  x方向の空間刻み幅
dy    :float = dx                      #[m]  y方向の空間刻み幅
nx    :int   = int((x_max - x_min) / dx)    #x方向の空間ステップ数
ny    :int   = int((y_max - y_min) / dy)    #y方向の空間ステップ数
x :np.ndarray = np.arange(x_min,x_max,dx)   #x方向の格子点座標の配列
y :np.ndarray = np.arange(y_min,y_max,dy)   #y方向の格子点座標の配列

#計算領域(時間)
t_min :float = settings_i['G13'].value #[s] 計算開始時刻(シミュレーション内部)
t_max :float = settings_i['G14'].value #[s] 計算終了時刻(シミュレーション内部)
dt    :float = settings_i['G15'].value #[s] 時間刻み幅
nt    :int   = int((t_max - t_min) / dt)   #時間ステップ数
t     :float = t_min                   #[s] シミュレーション内時刻
spf   :float = 1/30 if animation_speed == 'normal' else dt #アニメーションのフレームレート
spf          = int(spf / dt) * dt          #撮像スピードをdtの整数倍に調整

#境界条件
x_bc  :str = settings_i['G18'].value #計算領域の左右の境界条件
y_bc  :str = settings_i['G19'].value #計算領域の上下の境界条件

#初期条件
barrier    :np.ndarray = get_list_2d(baSheet_i,1,nx,1,ny)
if (drawing_mode == 'dot'         or 
    drawing_mode == 'past line'   or
    drawing_mode == 'stream line' or
    drawing_mode == 'streak line'   ):
    marker_scr :np.ndarray = get_list_2d(mkSheet_i,1,nx,1,ny)
else:
    marker_scr :np.ndarray = np.zeros((nx, ny))
ux         :np.ndarray = get_list_2d(uxSheet_i,1,nx,1,ny)
uy         :np.ndarray = get_list_2d(uySheet_i,1,nx,1,ny)
p          :np.ndarray = get_list_2d( pSheet_i,1,nx,1,ny)
Kx         :np.ndarray = get_list_2d(KxSheet_i,1,nx,1,ny) * (np.ones((nx, ny)) - barrier)
Ky         :np.ndarray = get_list_2d(KySheet_i,1,nx,1,ny) * (np.ones((nx, ny)) - barrier)

if   (drawing_mode == 'dot'         or
      drawing_mode == 'past line'   or
      drawing_mode == 'stream line' or
      drawing_mode == 'streak line'   ):
       colormap = cm.bone
elif (drawing_mode == 'vorticity'     ):
       colormap = cm.seismic
else :
       colormap = cm.jet

#以下内部設定-----------------------------------------------------------------------------

marker_scr_first :np.ndarray = marker_scr.copy() #markerスクリーンの初期値を記録
n_psolve = 100         #圧力ソルバーの反復回数
b = np.zeros((nx, ny)) #圧力ソルバーで使うソース項

#表示するスクリーンの初期化
screen = np.zeros((nx, ny))

#壁と角(境界条件用)
w_l   :np.ndarray = np.zeros((nx, ny)) #左    壁  ^
w_r   :np.ndarray = np.zeros((nx, ny)) #右        |  lu *u ru
w_d   :np.ndarray = np.zeros((nx, ny)) #下        j  l* ** r*
w_u   :np.ndarray = np.zeros((nx, ny)) #上        |  ld *d rd
e_ld  :np.ndarray = np.zeros((nx, ny)) #左下  角     --  i -- >
e_lu  :np.ndarray = np.zeros((nx, ny)) #左上     
e_rd  :np.ndarray = np.zeros((nx, ny)) #右下
e_ru  :np.ndarray = np.zeros((nx, ny)) #右上
w_all :np.ndarray = np.zeros((nx, ny)) #全ての壁と角

#barrierから壁と角を抽出する(境界条件用)
for i in range(1,barrier.shape[0] - 1):
    for j in range(1,barrier.shape[1] - 1):
        if barrier[i,j] == 1:

            if  (barrier[i-1,j] == 0 
               and ((barrier[i,j-1] == 1 and barrier[i,j+1] == 1)
               or   (barrier[i,j-1] == 0 and barrier[i,j+1] == 0))):
               w_l[i,j] = 1
            elif(barrier[i+1,j] == 0 
               and ((barrier[i,j-1] == 1 and barrier[i,j+1] == 1)
               or   (barrier[i,j-1] == 0 and barrier[i,j+1] == 0))):
               w_r[i,j] = 1
            elif(barrier[i,j-1] == 0 
               and ((barrier[i-1,j] == 1 and barrier[i+1,j] == 1)
               or   (barrier[i-1,j] == 0 and barrier[i+1,j] == 0))):
                w_d[i,j] = 1
            elif(barrier[i,j+1] == 0 
               and ((barrier[i-1,j] == 1 and barrier[i+1,j] == 1)
               or   (barrier[i-1,j] == 0 and barrier[i+1,j] == 0))):
               w_u[i,j] = 1

            elif(barrier[i-1,j-1] == 0 
               and (barrier[i-1,j] == 0 and barrier[i,j-1] == 0)):
               e_ld[i,j] = 1
            elif(barrier[i-1,j+1] == 0 
               and (barrier[i-1,j] == 0 and barrier[i,j+1] == 0)):
               e_lu[i,j] = 1
            elif(barrier[i+1,j-1] == 0 
               and (barrier[i+1,j] == 0 and barrier[i,j-1] == 0)):
               e_rd[i,j] = 1
            elif(barrier[i+1,j+1] == 0 
               and (barrier[i+1,j] == 0 and barrier[i,j+1] == 0)):
               e_ru[i,j] = 1

w_all :np.ndarray = w_l + w_r + w_d + w_u + e_ld + e_lu + e_rd + e_ru

#マーカー
numOfmarker = int(nx * ny / 10) #ランダムマーカーの数
marker_x :list = []
marker_y :list = []

show_barrier = 0.7                    #barrierをスクリーンに映す濃さ
marker_scr   = barrier * show_barrier #マーカーの位置を表示するスクリーン

#解析結果を格納するフォルダ作成と設定ファイルの格納
dirName = start_time.strftime('result\\' + folder_name +'_' + '%Y_%m%d_%H%M%S') #作成するフォルダ名

if not os.path.exists('result'):
    os.mkdir('result')
os.mkdir(dirName)                                        #フォルダ作成
os.mkdir(dirName + '\\01_First')                #初期条件フォルダ作成
os.mkdir(dirName + '\\02_Process')              #アニメフォルダ作成
os.mkdir(dirName + '\\03_Last')                 #終期条件フォルダ作成
sh.copy('config.xlsx',dirName + '\\01_First')   #初期条件フォルダに初期条件を格納
sh.copy('config.xlsx',dirName + '\\03_Last' )   #終期条件フォルダに初期条件を格納
book_o :xl.Workbook = xl.load_workbook(dirName + '\\03_Last\\config.xlsx', data_only=True)
settings_o = book_o['settings']
mkSheet_o  = book_o['marker'  ]
uxSheet_o  = book_o['ux'      ]
uySheet_o  = book_o['uy'      ]
pSheet_o   = book_o['p'       ]

result_all = 'config.xlsx 読込完了'
print(result_all)